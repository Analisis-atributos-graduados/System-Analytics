import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

class NumberedCanvas(canvas.Canvas):
    """
    Custom canvas to enable 2-pass page numbering ('Página X de Y')
    and header running titles.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8.5)
        self.setFillColor(colors.HexColor("#4A5568"))
        
        if self._pageNumber > 1:
            self.drawString(54, 750, "EvalIA - Reporte Estadístico Académico")
            self.setStrokeColor(colors.HexColor("#CBD5E1"))
            self.setLineWidth(0.5)
            self.line(54, 742, 558, 742)
            
        page_text = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(558, 40, page_text)
        self.drawString(54, 40, "EvalIA - Sistema de Evaluación Académica Inteligente")
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(54, 52, 558, 52)
        
        self.restoreState()

class ReportService:
    def _setup_styles(self):
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#1A365D"),
            spaceAfter=15
        )
        h1_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=13,
            leading=17,
            textColor=colors.HexColor("#2C3E50"),
            spaceBefore=14,
            spaceAfter=8,
            keepWithNext=True
        )
        h2_style = ParagraphStyle(
            'SubsectionHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=11,
            leading=15,
            textColor=colors.HexColor("#4A5568"),
            spaceBefore=10,
            spaceAfter=6,
            keepWithNext=True
        )
        body_style = ParagraphStyle(
            'BodyTextCustom',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9.5,
            leading=13.5,
            textColor=colors.HexColor("#2D3748"),
            spaceAfter=6
        )
        table_text_style = ParagraphStyle(
            'TableText',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8.5,
            leading=11.5,
            textColor=colors.HexColor("#2D3748")
        )
        table_text_bold_style = ParagraphStyle(
            'TableTextBold',
            parent=table_text_style,
            fontName='Helvetica-Bold'
        )
        table_header_style = ParagraphStyle(
            'TableHeaderCustom',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8.5,
            leading=11.5,
            textColor=colors.white
        )
        metadata_label_style = ParagraphStyle(
            'MetaLabel',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9.5,
            leading=13.5,
            textColor=colors.HexColor("#1A365D")
        )
        metadata_val_style = ParagraphStyle(
            'MetaValue',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9.5,
            leading=13.5,
            textColor=colors.HexColor("#2D3748")
        )
        
        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1A365D")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#F8F9FA"), colors.HexColor("#FFFFFF")]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
        ])
        
        return (title_style, h1_style, h2_style, body_style, table_text_style, table_text_bold_style, table_header_style, metadata_label_style, metadata_val_style, t_style)

    def generate_professor_report_pdf(self, stats: dict, metadata: dict) -> io.BytesIO:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=54,
            rightMargin=54,
            topMargin=54,
            bottomMargin=54
        )
        
        styles = self._setup_styles()
        title_style, h1_style, h2_style, body_style, table_text_style, table_text_bold_style, table_header_style, metadata_label_style, metadata_val_style, t_style = styles
        
        story = []
        
        story.append(Paragraph("Reporte Estadístico Académico de Evaluación", title_style))
        story.append(Spacer(1, 15))
        
        meta_table_data = [
            [Paragraph("Curso:", metadata_label_style), Paragraph(metadata.get('curso', 'N/A'), metadata_val_style),
             Paragraph("Semestre / Ciclo:", metadata_label_style), Paragraph(metadata.get('semestre', 'N/A'), metadata_val_style)],
            [Paragraph("Profesor:", metadata_label_style), Paragraph(metadata.get('profesor', 'N/A'), metadata_val_style),
             Paragraph("Fecha de Emisión:", metadata_label_style), Paragraph(datetime.now().strftime("%d/%m/%Y"), metadata_val_style)],
            [Paragraph("Tema Evaluado:", metadata_label_style), Paragraph(metadata.get('tema', 'N/A'), metadata_val_style),
             Paragraph("", metadata_label_style), Paragraph("", metadata_val_style)]
        ]
        meta_table = Table(meta_table_data, colWidths=[80, 172, 100, 152])
        meta_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("1. Resumen de Rendimiento", h1_style))
        total = stats['general']['total']
        summary_table_data = [
            [Paragraph("Métrica", table_header_style), Paragraph("Valor", table_header_style)],
            [Paragraph("Total de Alumnos Evaluados", table_text_style), Paragraph(str(total), table_text_style)],
            [Paragraph("Nota Promedio General", table_text_style), Paragraph(str(stats['general']['promedio']), table_text_style)],
            [Paragraph("Alumnos Aprobados (Nota >= 10.5)", table_text_style), Paragraph(f"{stats['general']['aprobados']} ({round(stats['general']['aprobados']/total*100, 1) if total > 0 else 0}%)", table_text_style)],
            [Paragraph("Alumnos Desaprobados", table_text_style), Paragraph(f"{stats['general']['desaprobados']} ({round(stats['general']['desaprobados']/total*100, 1) if total > 0 else 0}%)", table_text_style)]
        ]
        summary_table = Table(summary_table_data, colWidths=[240, 264])
        summary_table.setStyle(t_style)
        story.append(summary_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("2. Distribución de Notas", h1_style))
        dist_table_data = [
            [Paragraph("Rango de Notas", table_header_style), Paragraph("Cantidad de Estudiantes", table_header_style), Paragraph("Porcentaje", table_header_style)]
        ]
        for range_lbl, count in stats['distribucion'].items():
            pct = round(count / total * 100, 1) if total > 0 else 0
            dist_table_data.append([
                Paragraph(range_lbl, table_text_style),
                Paragraph(str(count), table_text_style),
                Paragraph(f"{pct}%", table_text_style)
            ])
        dist_table = Table(dist_table_data, colWidths=[150, 170, 184])
        dist_table.setStyle(t_style)
        story.append(dist_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("3. Promedio de Notas por Criterio", h1_style))
        crit_table_data = [
            [Paragraph("Criterio de Evaluación", table_header_style), Paragraph("Nota Promedio", table_header_style), Paragraph("Porcentaje de Logro", table_header_style)]
        ]
        for c in stats['criterios']:
            crit_table_data.append([
                Paragraph(c['nombre'], table_text_style),
                Paragraph(str(c['promedio']), table_text_style),
                Paragraph(f"{c['porcentaje']}%", table_text_style)
            ])
        crit_table = Table(crit_table_data, colWidths=[260, 120, 124])
        crit_table.setStyle(t_style)
        story.append(crit_table)
        story.append(Spacer(1, 15))
        
        story.append(PageBreak())
        story.append(Paragraph("4. Detalle de Calificaciones por Estudiante", h1_style))
        
        story.append(Paragraph("<b>Leyenda de Criterios de Rúbrica:</b>", table_text_bold_style))
        story.append(Spacer(1, 4))
        legend_data = [
            [Paragraph("Código", table_header_style), Paragraph("Nombre del Criterio", table_header_style)]
        ]
        for idx, c in enumerate(stats['criterios'], 1):
            legend_data.append([
                Paragraph(f"C{idx}", table_text_bold_style),
                Paragraph(c['nombre'], table_text_style)
            ])
        legend_table = Table(legend_data, colWidths=[60, 444])
        legend_table.setStyle(t_style)
        story.append(legend_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("<b>Calificaciones:</b>", table_text_bold_style))
        story.append(Spacer(1, 4))
        
        crit_headers = [f"C{i}" for i in range(1, len(stats['criterios']) + 1)]
        headers = [Paragraph("Estudiante", table_header_style), Paragraph("Nota", table_header_style)] + [Paragraph(h, table_header_style) for h in crit_headers]
        
        stud_table_data = [headers]
        criteria_names = [c['nombre'] for c in stats['criterios']]
        
        for student in stats['estudiantes']:
            row = [
                Paragraph(student['nombre'], table_text_style),
                Paragraph(str(student['nota']), table_text_style)
            ]
            for c_name in criteria_names:
                score_val = "-"
                for sc in student.get('criterios', []):
                    if sc['nombre'] == c_name:
                        score_val = str(sc.get('puntaje', '-'))
                        break
                row.append(Paragraph(score_val, table_text_style))
            stud_table_data.append(row)
            
        num_crit = len(stats['criterios'])
        col_widths = [200, 50] + [254 / num_crit if num_crit > 0 else 254] * num_crit
        
        stud_table = Table(stud_table_data, colWidths=col_widths)
        stud_table.setStyle(t_style)
        story.append(stud_table)
        story.append(Spacer(1, 15))
        
        fb = stats.get('feedback_global') or {}
        has_fb = fb.get('hallazgos') or fb.get('fortalezas') or fb.get('oportunidades')
        
        if has_fb:
            story.append(PageBreak())
            story.append(Paragraph("5. Resultado Global de la Evaluación", h1_style))
            
            if fb.get('hallazgos'):
                story.append(Paragraph("Hallazgos generales:", h2_style))
                story.append(Paragraph(fb['hallazgos'].replace("\n", "<br/>"), body_style))
                story.append(Spacer(1, 10))
                
            if fb.get('fortalezas'):
                story.append(Paragraph("Fortalezas demostradas por los estudiantes:", h2_style))
                story.append(Paragraph(fb['fortalezas'].replace("\n", "<br/>"), body_style))
                story.append(Spacer(1, 10))
                
            if fb.get('oportunidades'):
                story.append(Paragraph("Oportunidades de mejora identificadas:", h2_style))
                story.append(Paragraph(fb['oportunidades'].replace("\n", "<br/>"), body_style))
                story.append(Spacer(1, 10))
                
        doc.build(story, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        return buffer

    def generate_quality_report_pdf(self, stats: dict, metadata: dict) -> io.BytesIO:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=54,
            rightMargin=54,
            topMargin=54,
            bottomMargin=54
        )
        
        styles = self._setup_styles()
        title_style, h1_style, h2_style, body_style, table_text_style, table_text_bold_style, table_header_style, metadata_label_style, metadata_val_style, t_style = styles
        
        story = []
        
        story.append(Paragraph("Reporte de Calidad Educativa por Atributo", title_style))
        story.append(Spacer(1, 15))
        
        meta_table_data = [
            [Paragraph("Ciclo / Semestre:", metadata_label_style), Paragraph(metadata.get('semestre', 'N/A'), metadata_val_style),
             Paragraph("Atributo de Graduado:", metadata_label_style), Paragraph(metadata.get('atributo', 'N/A'), metadata_val_style)],
            [Paragraph("Curso:", metadata_label_style), Paragraph(metadata.get('curso', 'Todos los cursos'), metadata_val_style),
             Paragraph("Fecha de Generación:", metadata_label_style), Paragraph(datetime.now().strftime("%d/%m/%Y"), metadata_val_style)]
        ]
        
        fac = metadata.get('facultad')
        esc = metadata.get('escuela')
        nrc = metadata.get('nrc')
        
        if fac or esc or nrc:
            meta_table_data.append([
                Paragraph("Facultad:", metadata_label_style), Paragraph(fac or "Todas", metadata_val_style),
                Paragraph("Escuela:", metadata_label_style), Paragraph(esc or "Todas", metadata_val_style)
            ])
            if nrc:
                meta_table_data.append([
                    Paragraph("NRC:", metadata_label_style), Paragraph(str(nrc), metadata_val_style),
                    Paragraph("", metadata_label_style), Paragraph("", metadata_val_style)
                ])
                
        meta_table = Table(meta_table_data, colWidths=[100, 152, 120, 132])
        meta_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("1. Consolidado de Rendimiento", h1_style))
        summary_table_data = [
            [Paragraph("Métrica", table_header_style), Paragraph("Valor", table_header_style)],
            [Paragraph("Total de Alumnos Evaluados", table_text_style), Paragraph(str(stats.get('total_alumnos', 0)), table_text_style)],
            [Paragraph("Porcentaje de Logro (Excelente / Bueno)", table_text_style), Paragraph(f"{stats.get('porcentaje_logro', 0)}%", table_text_style)]
        ]
        summary_table = Table(summary_table_data, colWidths=[240, 264])
        summary_table.setStyle(t_style)
        story.append(summary_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("2. Distribución de Desempeño por Niveles", h1_style))
        
        criterio = stats.get('criterios', [{}])[0] if stats.get('criterios') else {}
        total = stats.get('total_alumnos', 0)
        
        excelente = criterio.get('excelente', 0)
        bueno = criterio.get('bueno', 0)
        mejora = criterio.get('requiereMejora', 0)
        no_aceptable = criterio.get('noAceptable', 0)
        
        pct_exc = round(excelente / total * 100, 1) if total > 0 else 0
        pct_bue = round(bueno / total * 100, 1) if total > 0 else 0
        pct_mej = round(mejora / total * 100, 1) if total > 0 else 0
        pct_nac = round(no_aceptable / total * 100, 1) if total > 0 else 0
        
        dist_table_data = [
            [Paragraph("Nivel de Desempeño", table_header_style), Paragraph("Rango de Nota", table_header_style), Paragraph("Cantidad de Alumnos", table_header_style), Paragraph("Porcentaje", table_header_style)],
            [Paragraph("Excelente", table_text_style), Paragraph("16 - 20", table_text_style), Paragraph(str(excelente), table_text_style), Paragraph(f"{pct_exc}%", table_text_style)],
            [Paragraph("Bueno", table_text_style), Paragraph("11 - 15", table_text_style), Paragraph(str(bueno), table_text_style), Paragraph(f"{pct_bue}%", table_text_style)],
            [Paragraph("Requiere Mejora", table_text_style), Paragraph("06 - 10", table_text_style), Paragraph(str(mejora), table_text_style), Paragraph(f"{pct_mej}%", table_text_style)],
            [Paragraph("No Aceptable", table_text_style), Paragraph("00 - 05", table_text_style), Paragraph(str(no_aceptable), table_text_style), Paragraph(f"{pct_nac}%", table_text_style)],
        ]
        
        dist_table = Table(dist_table_data, colWidths=[150, 100, 130, 124])
        dist_table.setStyle(t_style)
        story.append(dist_table)
        story.append(Spacer(1, 15))
        
        feedbacks = stats.get('feedbacks_profesores', [])
        if feedbacks:
            story.append(PageBreak())
            story.append(Paragraph("3. Observaciones y Resultados por Docente", h1_style))
            
            for fb in feedbacks:
                prof_story = []
                prof_story.append(Paragraph(f"👨‍🏫 Docente: {fb.get('profesor', 'Desconocido')} | Tema: {fb.get('tema', 'N/A')}", h2_style))
                prof_story.append(Spacer(1, 4))
                
                prof_story.append(Paragraph("<b>Hallazgos:</b>", table_text_bold_style))
                prof_story.append(Paragraph((fb.get('hallazgos') or "No registrados").replace("\n", "<br/>"), body_style))
                prof_story.append(Spacer(1, 6))
                
                prof_story.append(Paragraph("<b>Fortalezas logradas por los estudiantes:</b>", table_text_bold_style))
                prof_story.append(Paragraph((fb.get('fortalezas') or "Ninguna registrada").replace("\n", "<br/>"), body_style))
                prof_story.append(Spacer(1, 6))
                
                prof_story.append(Paragraph("<b>Oportunidades de mejora identificadas:</b>", table_text_bold_style))
                prof_story.append(Paragraph((fb.get('oportunidades') or "Ninguna registrada").replace("\n", "<br/>"), body_style))
                prof_story.append(Spacer(1, 15))
                
                story.append(KeepTogether(prof_story))
                story.append(Spacer(1, 10))
                
        doc.build(story, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        return buffer

    def generate_professor_report_excel(self, stats: dict, metadata: dict) -> io.BytesIO:
        wb = openpyxl.Workbook()
        
        font_title = Font(name="Calibri", size=15, bold=True, color="1A365D")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_section = Font(name="Calibri", size=12, bold=True, color="1A365D")
        font_bold = Font(name="Calibri", size=10, bold=True)
        font_regular = Font(name="Calibri", size=10)
        
        fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        ws1 = wb.active
        ws1.title = "Resumen"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.append(["Reporte Estadístico Académico de Evaluación"])
        ws1.cell(row=1, column=1).font = font_title
        ws1.row_dimensions[1].height = 25
        ws1.append([])
        
        ws1.append(["Curso:", metadata.get('curso', 'N/A'), "", "Semestre:", metadata.get('semestre', 'N/A')])
        ws1.append(["Profesor:", metadata.get('profesor', 'N/A'), "", "Fecha:", datetime.now().strftime("%d/%m/%Y")])
        ws1.append(["Tema:", metadata.get('tema', 'N/A')])
        
        for r in range(3, 6):
            ws1.cell(row=r, column=1).font = font_bold
            ws1.cell(row=r, column=4).font = font_bold
            for col in [1, 2, 4, 5]:
                ws1.cell(row=r, column=col).border = thin_border
        ws1.append([])
        
        ws1.append(["1. Resumen de Rendimiento"])
        ws1.cell(row=7, column=1).font = font_section
        ws1.append(["Métrica", "Valor"])
        
        total = stats['general']['total']
        summary_rows = [
            ["Total de Alumnos Evaluados", total],
            ["Nota Promedio General", stats['general']['promedio']],
            ["Alumnos Aprobados (Nota >= 10.5)", f"{stats['general']['aprobados']} ({round(stats['general']['aprobados']/total*100, 1) if total > 0 else 0}%)"],
            ["Alumnos Desaprobados", f"{stats['general']['desaprobados']} ({round(stats['general']['desaprobados']/total*100, 1) if total > 0 else 0}%)"]
        ]
        
        for row in summary_rows:
            ws1.append(row)
            
        for col in [1, 2]:
            c = ws1.cell(row=8, column=col)
            c.font = font_header
            c.fill = fill_header
            c.border = thin_border
            
        for r_idx in range(9, 13):
            fill = fill_zebra if r_idx % 2 == 1 else fill_white
            for col in [1, 2]:
                c = ws1.cell(row=r_idx, column=col)
                c.font = font_regular
                c.fill = fill
                c.border = thin_border
                
        ws1.append([])
        
        ws1.append(["2. Distribución de Notas"])
        ws1.cell(row=14, column=1).font = font_section
        ws1.append(["Rango de Notas", "Cantidad", "Porcentaje"])
        
        dist_start_row = 16
        for range_lbl, count in stats['distribucion'].items():
            pct = count / total if total > 0 else 0
            ws1.append([range_lbl, count, pct])
            
        for col in [1, 2, 3]:
            c = ws1.cell(row=15, column=col)
            c.font = font_header
            c.fill = fill_header
            c.border = thin_border
            
        num_dist_rows = len(stats['distribucion'])
        for r_idx in range(dist_start_row, dist_start_row + num_dist_rows):
            fill = fill_zebra if r_idx % 2 == 1 else fill_white
            for col in [1, 2, 3]:
                c = ws1.cell(row=r_idx, column=col)
                c.fill = fill
                c.border = thin_border
                if col == 3:
                    c.number_format = '0.0%'
                    c.font = font_regular
                else:
                    c.font = font_regular
                    
        ws1.append([])
        
        ws1.append(["3. Promedio por Criterio"])
        ws1.cell(row=15 + num_dist_rows + 2, column=1).font = font_section
        ws1.append(["Criterio", "Promedio", "Porcentaje de Logro"])
        
        crit_start_row = 15 + num_dist_rows + 4
        for c in stats['criterios']:
            ws1.append([c['nombre'], c['promedio'], c['porcentaje']/100])
            
        for col in [1, 2, 3]:
            c = ws1.cell(row=crit_start_row - 1, column=col)
            c.font = font_header
            c.fill = fill_header
            c.border = thin_border
            
        num_crit = len(stats['criterios'])
        for r_idx in range(crit_start_row, crit_start_row + num_crit):
            fill = fill_zebra if r_idx % 2 == 1 else fill_white
            for col in [1, 2, 3]:
                c = ws1.cell(row=r_idx, column=col)
                c.fill = fill
                c.border = thin_border
                if col == 3:
                    c.number_format = '0.0%'
                    c.font = font_regular
                else:
                    c.font = font_regular
                    
        ws1.append([])
        
        obs_start = crit_start_row + num_crit + 1
        ws1.append(["4. Observaciones de la Evaluación"])
        ws1.cell(row=obs_start, column=1).font = font_section
        
        fb = stats.get('feedback_global') or {}
        ws1.append(["Módulo", "Detalle"])
        
        ws1.append(["Hallazgos generales", fb.get('hallazgos') or "No registrados"])
        ws1.append(["Fortalezas logradas", fb.get('fortalezas') or "Ninguna registrada"])
        ws1.append(["Oportunidades de mejora", fb.get('oportunidades') or "Ninguna registrada"])
        
        for col in [1, 2]:
            c = ws1.cell(row=obs_start + 1, column=col)
            c.font = font_header
            c.fill = fill_header
            c.border = thin_border
            
        for r_idx in range(obs_start + 2, obs_start + 5):
            for col in [1, 2]:
                c = ws1.cell(row=r_idx, column=col)
                c.font = font_regular
                c.border = thin_border
                if col == 2:
                    c.alignment = Alignment(wrap_text=True)
                    
        for col in ws1.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if '\n' in val:
                    val = max(val.split('\n'), key=len)
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        ws2 = wb.create_sheet(title="Estudiantes")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2.append(["Detalle de Calificaciones por Estudiante"])
        ws2.cell(row=1, column=1).font = font_title
        ws2.append([])
        
        ws2.append(["Leyenda de Criterios:"])
        ws2.cell(row=3, column=1).font = font_bold
        ws2.append(["Código", "Nombre del Criterio"])
        
        for idx, c in enumerate(stats['criterios'], 1):
            ws2.append([f"C{idx}", c['nombre']])
            
        for col in [1, 2]:
            c = ws2.cell(row=4, column=col)
            c.font = font_header
            c.fill = fill_header
            c.border = thin_border
            
        for r_idx in range(5, 5 + num_crit):
            fill = fill_zebra if r_idx % 2 == 1 else fill_white
            for col in [1, 2]:
                c = ws2.cell(row=r_idx, column=col)
                c.font = font_regular
                c.fill = fill
                c.border = thin_border
                
        ws2.append([])
        
        grades_start = 5 + num_crit + 2
        ws2.append(["Calificaciones:"])
        ws2.cell(row=grades_start, column=1).font = font_bold
        
        crit_headers = [f"C{i}" for i in range(1, len(stats['criterios']) + 1)]
        headers = ["Estudiante", "Nota Final"] + crit_headers
        ws2.append(headers)
        
        criteria_names = [c['nombre'] for c in stats['criterios']]
        for student in stats['estudiantes']:
            row = [student['nombre'], student['nota']]
            for c_name in criteria_names:
                score_val = "-"
                for sc in student.get('criterios', []):
                    if sc['nombre'] == c_name:
                        score_val = sc.get('puntaje', '-')
                        break
                row.append(score_val)
            ws2.append(row)
            
        num_students = len(stats['estudiantes'])
        
        for col_idx in range(1, len(headers) + 1):
            c = ws2.cell(row=grades_start + 1, column=col_idx)
            c.font = font_header
            c.fill = fill_header
            c.border = thin_border
            
        for r_idx in range(grades_start + 2, grades_start + 2 + num_students):
            fill = fill_zebra if r_idx % 2 == 1 else fill_white
            for col_idx in range(1, len(headers) + 1):
                c = ws2.cell(row=r_idx, column=col_idx)
                c.font = font_regular
                c.fill = fill
                c.border = thin_border
                
        for col in ws2.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if '\n' in val:
                    val = max(val.split('\n'), key=len)
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    def generate_quality_report_excel(self, stats: dict, metadata: dict) -> io.BytesIO:
        wb = openpyxl.Workbook()
        
        font_title = Font(name="Calibri", size=15, bold=True, color="1A365D")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_section = Font(name="Calibri", size=12, bold=True, color="1A365D")
        font_bold = Font(name="Calibri", size=10, bold=True)
        font_regular = Font(name="Calibri", size=10)
        
        fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        ws1 = wb.active
        ws1.title = "Consolidado"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.append(["Reporte de Calidad Educativa por Atributo"])
        ws1.cell(row=1, column=1).font = font_title
        ws1.row_dimensions[1].height = 25
        ws1.append([])
        
        ws1.append(["Ciclo / Semestre:", metadata.get('semestre', 'N/A'), "", "Atributo de Graduado:", metadata.get('atributo', 'N/A')])
        ws1.append(["Curso:", metadata.get('curso', 'Todos los cursos'), "", "Fecha de Generación:", datetime.now().strftime("%d/%m/%Y")])
        
        meta_row = 5
        fac = metadata.get('facultad')
        esc = metadata.get('escuela')
        nrc = metadata.get('nrc')
        
        if fac or esc or nrc:
            ws1.append(["Facultad:", fac or "Todas", "", "Escuela:", esc or "Todas"])
            meta_row += 1
            if nrc:
                ws1.append(["NRC:", str(nrc)])
                meta_row += 1
                
        for r in range(3, meta_row):
            ws1.cell(row=r, column=1).font = font_bold
            ws1.cell(row=r, column=4).font = font_bold
            for col in [1, 2, 4, 5]:
                ws1.cell(row=r, column=col).border = thin_border
                
        ws1.append([])
        
        sect1_row = meta_row + 1
        ws1.append(["1. Consolidado de Rendimiento"])
        ws1.cell(row=sect1_row, column=1).font = font_section
        ws1.append(["Métrica", "Valor"])
        
        summary_rows = [
            ["Total de Alumnos Evaluados", stats.get('total_alumnos', 0)],
            ["Porcentaje de Logro (Excelente / Bueno)", f"{stats.get('porcentaje_logro', 0)}%"]
        ]
        
        for row in summary_rows:
            ws1.append(row)
            
        for col in [1, 2]:
            c = ws1.cell(row=sect1_row + 1, column=col)
            c.font = font_header
            c.fill = fill_header
            c.border = thin_border
            
        for r_idx in range(sect1_row + 2, sect1_row + 4):
            for col in [1, 2]:
                c = ws1.cell(row=r_idx, column=col)
                c.font = font_regular
                c.border = thin_border
                
        ws1.append([])
        
        sect2_row = sect1_row + 5
        ws1.append(["2. Distribución de Desempeño por Niveles"])
        ws1.cell(row=sect2_row, column=1).font = font_section
        ws1.append(["Nivel de Desempeño", "Rango de Nota", "Cantidad", "Porcentaje"])
        
        criterio = stats.get('criterios', [{}])[0] if stats.get('criterios') else {}
        total = stats.get('total_alumnos', 0)
        
        excelente = criterio.get('excelente', 0)
        bueno = criterio.get('bueno', 0)
        mejora = criterio.get('requiereMejora', 0)
        no_aceptable = criterio.get('noAceptable', 0)
        
        pct_exc = excelente / total if total > 0 else 0
        pct_bue = bueno / total if total > 0 else 0
        pct_mej = mejora / total if total > 0 else 0
        pct_nac = no_aceptable / total if total > 0 else 0
        
        ws1.append(["Excelente", "16 - 20", excelente, pct_exc])
        ws1.append(["Bueno", "11 - 15", bueno, pct_bue])
        ws1.append(["Requiere Mejora", "06 - 10", mejora, pct_mej])
        ws1.append(["No Aceptable", "00 - 05", no_aceptable, pct_nac])
        
        for col in [1, 2, 3, 4]:
            c = ws1.cell(row=sect2_row + 1, column=col)
            c.font = font_header
            c.fill = fill_header
            c.border = thin_border
            
        for r_idx in range(sect2_row + 2, sect2_row + 6):
            fill = fill_zebra if r_idx % 2 == 1 else fill_white
            for col in [1, 2, 3, 4]:
                c = ws1.cell(row=r_idx, column=col)
                c.fill = fill
                c.border = thin_border
                if col == 4:
                    c.number_format = '0.0%'
                    c.font = font_regular
                else:
                    c.font = font_regular
                    
        for col in ws1.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if '\n' in val:
                    val = max(val.split('\n'), key=len)
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        feedbacks = stats.get('feedbacks_profesores', [])
        if feedbacks:
            ws2 = wb.create_sheet(title="Observaciones Docentes")
            ws2.views.sheetView[0].showGridLines = True
            
            ws2.append(["Resultados y Observaciones por Profesor"])
            ws2.cell(row=1, column=1).font = font_title
            ws2.append([])
            
            ws2.append(["Profesor", "Tema", "Hallazgos", "Fortalezas", "Oportunidades de Mejora"])
            
            for fb in feedbacks:
                ws2.append([
                    fb.get('profesor', 'Desconocido'),
                    fb.get('tema', 'N/A'),
                    fb.get('hallazgos') or "No registrados",
                    fb.get('fortalezas') or "Ninguna registrada",
                    fb.get('oportunidades') or "Ninguna registrada"
                ])
                
            num_fbs = len(feedbacks)
            
            for col_idx in range(1, 6):
                c = ws2.cell(row=3, column=col_idx)
                c.font = font_header
                c.fill = fill_header
                c.border = thin_border
                
            for r_idx in range(4, 4 + num_fbs):
                fill = fill_zebra if r_idx % 2 == 1 else fill_white
                for col_idx in range(1, 6):
                    c = ws2.cell(row=r_idx, column=col_idx)
                    c.font = font_regular
                    c.fill = fill
                    c.border = thin_border
                    if col_idx in [3, 4, 5]:
                        c.alignment = Alignment(wrap_text=True, vertical='top')
                    else:
                        c.alignment = Alignment(vertical='top')
                        
            for col in ws2.columns:
                max_len = 0
                for cell in col:
                    val = str(cell.value or '')
                    if '\n' in val:
                        val = max(val.split('\n'), key=len)
                    if len(val) > max_len:
                        max_len = len(val)
                col_letter = get_column_letter(col[0].column)
                if col[0].column in [3, 4, 5]:
                    ws2.column_dimensions[col_letter].width = 45
                else:
                    ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 30)
                    
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out
